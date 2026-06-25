#!/usr/bin/env python3
"""Extract Tier-1 note-hardware design files into clean markdown for the ragpi RAG index.

Tier 1 = the cleanly-answerable hardware content:
  * BOMs            (.xlsx / .csv)        -> markdown tables
  * KiCad schematics (.kicad_sch)         -> component + net summary (parsed s-expr)
  * Schematic PDFs  (.pdf, schematic only) -> per-page descriptions via a vision model (VLM)

Output is written under tools/extract_for_rag/rag-extracted/, mirroring the source tree, one
<source>.md per file. A future ragpi `github_readme` source points at that dir; creating that
source is out of scope for this tool (see tools/extract_for_rag/README.md).

Selection rules:
  * `_Legacy Hardware/` and non-hardware top-level dirs are excluded.
  * Version dedupe is **per (product, doc-type)**: for each product and each doc-type we keep only
    the files from the newest `vX.Y` folder that actually contains that doc-type. This avoids
    indexing the same board twice while preserving coverage when a newer rev only adds an
    unrelated document (e.g. Notecarrier-A's schematic/BOM/KiCad live in v2.0; v2.3 only adds a
    PCB doc, so we keep v2.0 for those types).

Usage (incremental is the default; --full is an explicit clean rebuild):
  python extract.py --changed-from FILE     # default mode: re-extract changed/missing, prune orphans
  python extract.py                         # incremental with no diff: generate missing + prune only
  python extract.py --full                  # clean rebuild: wipe + regenerate everything
  python extract.py --full --no-vlm         # rebuild BOM/KiCad only, skip costly schematic VLM
"""

from __future__ import annotations

import argparse
import base64
import csv
import os
import re
import sys
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path

# tools/extract_for_rag/extract.py -> repo root is two parents up.
REPO_ROOT = Path(__file__).resolve().parents[2]
TOOL_DIR = Path(__file__).resolve().parent
DEFAULT_OUT = TOOL_DIR / "rag-extracted"

# Top-level directories that are never scanned.
EXCLUDE_TOP = {
    ".git",
    ".github",
    "tools",                     # holds this pipeline and its rag-extracted/ output
    "_Legacy Hardware",          # discontinued boards, excluded per review
    "3D Models for Enclosures",  # 3D geometry only, no Tier-1 text
}

VERSION_RE = re.compile(r"^v(\d+(?:\.\d+)*)$", re.IGNORECASE)
# Abbreviated "sch" in a filename, e.g. "..._sch.PDF" or "..._SCH_RevA.pdf".
SCH_ABBR_RE = re.compile(r"(?:^|[_\-. ])sch(?:[_\-. ]|$)", re.IGNORECASE)
# Blues/Byte Lab schematic export convention: a 6-digit doc number + "Rev-NN",
# e.g. "100378_notecarrier-xm_Rev-11.PDF" (these are the multi-sheet schematics; their
# filenames contain no "sch"). Dimension (2200-*) and assembly (3000-*) docs don't match.
SCH_DOCNUM_RE = re.compile(r"^\d{6}[_-].*rev[-_ ]?\d", re.IGNORECASE)

# PDF path fragments that mean "not a schematic" even if "sch" appears nearby.
PDF_NEGATIVE = (
    "3d model", "assembly", "fabrication", "draftsman", "pcb drawing", "pcb_",
    "dimension", "dim-drawing", "dim_drawing", "board drawing", "panel",
    "block-diagram", "block diagram", "overview", "/reference/", "datasheet",
)

BOM, KICAD, SCHEMATIC = "bom", "kicad", "schematic"

# Path fragments (lowercase, "/"-joined) whose files are never indexed.
# KiCad_format/validation/ holds QA port-comparison artifacts (Gerber-diff PNGs and BOMs
# exported from both OrCAD/Altium and KiCad) that are ~90-93% replicants of the canonical BOM.
EXCLUDE_PATH = ("kicad_format/validation/",)

# A reference designator like C12, R5, U3, J2. Used for BOM dedup.
REFDES_RE = re.compile(r"^[A-Z]{1,3}\d+$")
# Two BOMs for the same board whose reference sets overlap at/above this Jaccard are duplicates.
BOM_DUP_JACCARD = 0.9

VLM_PROMPT = (
    "This is one page of an electronics hardware schematic. Transcribe it as clean markdown for a "
    "search index. List every component reference designator with its value/part, every connector "
    "and its pin labels, every named net/signal, and any notes in the title block. Use short "
    "headings and bullet lists. Do not invent details you cannot read; omit decorative borders."
)


@dataclass(frozen=True)
class Candidate:
    rel: Path          # path relative to REPO_ROOT
    product: str       # top-level product directory name
    doctype: str       # BOM | KICAD | SCHEMATIC
    version: tuple | None


# --------------------------------------------------------------------------- classification

def classify(rel: Path) -> str | None:
    """Return the Tier-1 doc-type for a source path, or None if it should not be indexed."""
    name = rel.name.lower()
    ext = rel.suffix.lower()
    path_l = "/".join(p.lower() for p in rel.parts)

    if any(frag in path_l for frag in EXCLUDE_PATH):
        return None

    if ext == ".kicad_sch":
        return KICAD

    if ext in (".xlsx", ".csv"):
        if "pick place" in path_l or name.startswith("pnp") or "p&p" in name:
            return None  # pick-and-place placement data, not a BOM
        if ext == ".xlsx":
            return BOM if ("bom" in name or "bill of material" in name) else None
        # Remaining CSVs are BOM / component-list / validation exports (OrCAD/KiCad) -> index.
        return BOM

    if ext == ".pdf":
        if any(neg in path_l for neg in PDF_NEGATIVE):
            return None
        if "schematic" in path_l or SCH_ABBR_RE.search(name) or SCH_DOCNUM_RE.search(name):
            return SCHEMATIC
        return None

    return None


def version_of(rel: Path) -> tuple | None:
    for part in rel.parts:
        m = VERSION_RE.match(part)
        if m:
            return tuple(int(x) for x in m.group(1).split("."))
    return None


def _vkey(v: tuple | None) -> tuple:
    return v if v is not None else (-1,)


def gather_candidates(root: Path) -> list[Candidate]:
    cands: list[Candidate] = []
    for top in sorted(p for p in root.iterdir() if p.is_dir() and p.name not in EXCLUDE_TOP):
        for f in top.rglob("*"):
            if not f.is_file() or ".git" in f.parts:
                continue
            rel = f.relative_to(root)
            dt = classify(rel)
            if dt is not None:
                cands.append(Candidate(rel, top.name, dt, version_of(rel)))
    return cands


def select_latest_per_doctype(cands: list[Candidate]) -> list[Candidate]:
    """Keep, for each (product, doctype), only files at the highest version present."""
    best: dict[tuple, tuple] = defaultdict(lambda: (-2,))
    for c in cands:
        key = (c.product, c.doctype)
        if _vkey(c.version) > best[key]:
            best[key] = _vkey(c.version)
    return [c for c in cands if _vkey(c.version) == best[(c.product, c.doctype)]]


def _refdes_set(path: Path) -> set:
    """Set of reference designators in a BOM file (best-effort; scans all cells)."""
    refs: set[str] = set()

    def add_token(tok: str):
        tok = tok.strip()
        m = re.match(r"^([A-Z]{1,3})(\d+)\s*-\s*(?:[A-Z]{1,3})?(\d+)$", tok)  # range, e.g. C4-C6
        if m:
            pre, a, b = m.group(1), int(m.group(2)), int(m.group(3))
            if 0 <= b - a < 500:
                refs.update(f"{pre}{i}" for i in range(a, b + 1))
                return
        if REFDES_RE.match(tok):
            refs.add(tok)

    def scan(cell):
        for piece in re.split(r"[,\s]+", str(cell)):
            add_token(piece)

    if path.suffix.lower() == ".csv":
        with open(path, newline="", encoding="utf-8", errors="replace") as fh:
            for row in csv.reader(fh):
                for cell in row:
                    if cell:
                        scan(cell)
    else:
        import openpyxl

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if cell is not None:
                        scan(cell)
        wb.close()
    return refs


def dedup_boms(selected: list[Candidate], root: Path, log_fn) -> list[Candidate]:
    """Drop a BOM when its reference set is ~identical to another already-kept BOM for the
    same product. Prefers .xlsx over .csv, then the shortest path. Distinct BOMs (e.g. Swan
    Carrier vs Feather) have low overlap and are kept."""
    by_product: dict[str, list[Candidate]] = defaultdict(list)
    for c in selected:
        if c.doctype == BOM:
            by_product[c.product].append(c)

    drop: set[str] = set()
    for boms in by_product.values():
        if len(boms) < 2:
            continue
        order = sorted(boms, key=lambda c: (c.rel.suffix.lower() != ".xlsx", len(c.rel.as_posix()), c.rel.as_posix()))
        kept: list[tuple[Candidate, set]] = []
        for c in order:
            try:
                refs = _refdes_set(root / c.rel)
            except Exception:  # noqa: BLE001 - if unreadable, keep it rather than guess
                kept.append((c, set()))
                continue
            dup_of = None
            for kc, krefs in kept:
                if refs and krefs:
                    jac = len(refs & krefs) / len(refs | krefs)
                    if jac >= BOM_DUP_JACCARD:
                        dup_of, jval = kc, jac
                        break
            if dup_of is not None:
                drop.add(c.rel.as_posix())
                log_fn(f"[extract] dedup BOM: drop '{c.rel.as_posix()}' (~{jval:.0%} same refs as '{dup_of.rel.as_posix()}')")
            else:
                kept.append((c, refs))
    return [c for c in selected if c.rel.as_posix() not in drop]


# --------------------------------------------------------------------------- markdown helpers

def _natkey(s: str):
    return [int(t) if t.isdigit() else t.lower() for t in re.split(r"(\d+)", s)]


def _cell(v) -> str:
    s = "" if v is None else str(v)
    return s.replace("|", r"\|").replace("\r", " ").replace("\n", " ").strip()


def rows_to_md_table(rows: list[list]) -> str:
    rows = [r for r in rows if any(_cell(c) for c in r)]
    if not rows:
        return "_(empty)_"
    width = max(len(r) for r in rows)
    header = [_cell(c) for c in rows[0]] + [""] * (width - len(rows[0]))
    header = [h or f"col{i+1}" for i, h in enumerate(header)]
    out = ["| " + " | ".join(header) + " |", "| " + " | ".join(["---"] * width) + " |"]
    for r in rows[1:]:
        cells = [_cell(c) for c in r] + [""] * (width - len(r))
        out.append("| " + " | ".join(cells) + " |")
    return "\n".join(out)


# --------------------------------------------------------------------------- extractors

def extract_csv(path: Path) -> str:
    with open(path, newline="", encoding="utf-8", errors="replace") as fh:
        rows = list(csv.reader(fh))
    return rows_to_md_table(rows)


def extract_xlsx(path: Path) -> str:
    import openpyxl  # local import: only the pipeline needs this dep

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    parts = []
    for ws in wb.worksheets:
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        table = rows_to_md_table(rows)
        if table != "_(empty)_":
            parts.append(f"### Sheet: {ws.title}\n\n{table}")
    wb.close()
    return "\n\n".join(parts) if parts else "_(no data)_"


def _parse_sexpr(text: str):
    tokens = re.findall(r'"(?:[^"\\]|\\.)*"|\(|\)|[^\s()]+', text)
    pos = 0

    def walk():
        nonlocal pos
        node = []
        while pos < len(tokens):
            t = tokens[pos]
            pos += 1
            if t == "(":
                node.append(walk())
            elif t == ")":
                return node
            else:
                node.append(t)
        return node

    return walk()


def _sval(tok):
    if isinstance(tok, str) and len(tok) >= 2 and tok.startswith('"') and tok.endswith('"'):
        return tok[1:-1]
    return tok


def _is_node(x, name) -> bool:
    return isinstance(x, list) and x and x[0] == name


def _props(symnode) -> dict:
    out = {}
    for c in symnode:
        if _is_node(c, "property"):
            out[_sval(c[1])] = _sval(c[2])
    return out


def extract_kicad(path: Path) -> str:
    parsed = _parse_sexpr(path.read_text(encoding="utf-8", errors="replace"))
    if not parsed or not isinstance(parsed[0], list):
        raise ValueError("empty or unparseable .kicad_sch")
    root = parsed[0]
    title = ""
    comps: list[tuple[str, str, str]] = []
    nets: set[str] = set()
    for c in root:
        if not isinstance(c, list):
            continue
        head = c[0]
        if head == "title_block":
            for t in c:
                if _is_node(t, "title"):
                    title = _sval(t[1])
        elif head == "lib_symbols":
            continue  # library definitions, not placed parts
        elif head == "symbol":
            p = _props(c)
            ref = p.get("Reference", "")
            if not ref or ref.startswith("#"):  # skip power/ground pseudo-symbols
                continue
            comps.append((ref, p.get("Value", ""), p.get("Footprint", "")))
        elif head in ("label", "global_label", "hierarchical_label"):
            if len(c) > 1:
                nets.add(_sval(c[1]))

    comps = sorted(set(comps), key=lambda r: _natkey(r[0]))
    md = []
    if title:
        md.append(f"_Sheet title: {title}_\n")
    md.append(f"### Components ({len(comps)})\n")
    md.append(rows_to_md_table([["Reference", "Value", "Footprint"], *comps]))
    if nets:
        md.append(f"\n### Nets / labels ({len(nets)})\n")
        md.append(", ".join(f"`{n}`" for n in sorted(nets, key=_natkey)))
    return "\n".join(md)


def extract_schematic_pdf(path: Path, vlm) -> tuple[str, int]:
    """Return (markdown, failed_page_count). A failed page yields a placeholder so one flaky page
    doesn't drop the whole schematic, but the count lets the caller surface it (and fail the run)
    rather than silently committing placeholder text."""
    import fitz  # pymupdf
    from concurrent.futures import ThreadPoolExecutor

    doc = fitz.open(path)
    pngs = [page.get_pixmap(dpi=200).tobytes("png") for page in doc]
    doc.close()

    def describe(i_png):
        i, png = i_png
        try:
            return i, vlm(png), False
        except Exception as e:  # noqa: BLE001 - one flaky page shouldn't drop the schematic
            return i, f"_(page {i + 1} extraction failed: {type(e).__name__}: {e})_", True

    workers = max(1, int(os.environ.get("RAG_VLM_CONCURRENCY", "6")))
    with ThreadPoolExecutor(max_workers=workers) as pool:
        results = list(pool.map(describe, enumerate(pngs)))
    texts = {i: text for i, text, _ in results}
    failed = sum(1 for _, _, f in results if f)
    body = "\n\n".join(f"## Page {i + 1}\n\n{texts[i]}" for i in range(len(pngs)))
    return body, failed


def make_vlm():
    """Return a function (png_bytes)->markdown, or None if VLM is unavailable."""
    if not os.environ.get("OPENAI_API_KEY"):
        return None
    from openai import OpenAI

    # Generous timeout + retries so transient rate-limits/timeouts don't count as page failures.
    client = OpenAI(timeout=180.0, max_retries=5)
    model = os.environ.get("RAG_VLM_MODEL", "gpt-5.5")

    def describe(png: bytes) -> str:
        b64 = base64.b64encode(png).decode()
        resp = client.chat.completions.create(
            model=model,
            messages=[{
                "role": "user",
                "content": [
                    {"type": "text", "text": VLM_PROMPT},
                    {"type": "image_url",
                     "image_url": {"url": f"data:image/png;base64,{b64}"}},
                ],
            }],
        )
        return (resp.choices[0].message.content or "").strip()

    return describe


# --------------------------------------------------------------------------- output

def _vstr(v: tuple | None) -> str:
    return "v" + ".".join(str(n) for n in v) if v else "-"


def write_markdown(out_dir: Path, c: Candidate, body: str) -> Path:
    dest = out_dir / c.rel.parent / (c.rel.name + ".md")
    dest.parent.mkdir(parents=True, exist_ok=True)
    front = (
        "---\n"
        f"product: {c.product}\n"
        f"version: {_vstr(c.version)}\n"
        f"doc_type: {c.doctype}\n"
        f"source_file: {c.rel.as_posix()}\n"
        "generated_by: tools/extract_for_rag/extract.py\n"
        "---\n\n"
        f"# {c.product} — {c.doctype.upper()} ({_vstr(c.version)})\n\n"
        f"_Source: `{c.rel.as_posix()}`_\n\n"
    )
    dest.write_text(front + body + "\n", encoding="utf-8")
    return dest


def _output_doctype(md: Path) -> str | None:
    """Infer the doc-type of a generated <source>.md from its source extension."""
    n = md.name.lower()
    if n.endswith(".kicad_sch.md"):
        return KICAD
    if n.endswith(".xlsx.md") or n.endswith(".csv.md"):
        return BOM
    if n.endswith(".pdf.md"):
        return SCHEMATIC
    return None


def clean_outputs(out_dir: Path, doctypes: set) -> int:
    """Delete existing generated markdown for the given doc-types (used before a full rebuild
    so regeneration is a clean slate — a file that fails to re-extract is absent, not stale)."""
    if not out_dir.exists():
        return 0
    removed = 0
    for md in out_dir.rglob("*.md"):
        if _output_doctype(md) in doctypes:
            md.unlink()
            removed += 1
    return removed


def prune_orphans(out_dir: Path, selected: list[Candidate]) -> int:
    keep = {(out_dir / c.rel.parent / (c.rel.name + ".md")).resolve() for c in selected}
    removed = 0
    for md in out_dir.rglob("*.md"):
        if md.resolve() not in keep:
            md.unlink()
            removed += 1
    # Remove now-empty directories left behind (deepest first).
    for d in sorted((p for p in out_dir.rglob("*") if p.is_dir()), key=lambda p: len(p.parts), reverse=True):
        if not any(d.iterdir()):
            d.rmdir()
    return removed


# --------------------------------------------------------------------------- main

def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--root", type=Path, default=REPO_ROOT, help="repo root (default: auto)")
    ap.add_argument("--out", type=Path, default=DEFAULT_OUT, help="output dir (default: tools/extract_for_rag/rag-extracted)")
    ap.add_argument("--full", action="store_true", help="force a clean rebuild (wipe + regenerate all); default is incremental")
    ap.add_argument("--changed-from", type=Path, help="incremental: file with newline-separated changed source paths (from `git diff --name-only`)")
    ap.add_argument("--no-vlm", action="store_true", help="skip schematic-PDF vision extraction")
    ap.add_argument("--limit", type=int, help="process at most N files (debug)")
    ap.add_argument("--only", choices=[BOM, KICAD, SCHEMATIC], help="restrict to one doc-type (debug)")
    args = ap.parse_args(argv)

    root: Path = args.root.resolve()
    out_dir: Path = args.out.resolve()

    cands = gather_candidates(root)
    selected = select_latest_per_doctype(cands)
    selected = dedup_boms(selected, root, lambda m: print(m, file=sys.stderr))

    # NOTE: --only filters what we (re)extract, NOT `selected`. prune_orphans reconciles against
    # the full selection, so a debug `--only schematic` run must not delete BOM/KiCad outputs.
    def _md_path(c: Candidate) -> Path:
        return out_dir / c.rel.parent / (c.rel.name + ".md")

    def _only_ok(c: Candidate) -> bool:
        return not args.only or c.doctype == args.only

    # Incremental is the default; --full is the explicit clean-rebuild opt-in.
    if args.full:
        changed = set()
        mode = "full"
    else:
        changed = set()
        if args.changed_from:
            changed = {
                line.strip()
                for line in args.changed_from.read_text(encoding="utf-8").splitlines()
                if line.strip()
            }
        mode = "incremental"

    vlm = None if args.no_vlm else make_vlm()
    vlm_available = vlm is not None

    if mode == "full":
        # Clean slate: delete existing outputs for the doc-types we're about to regenerate BEFORE
        # extracting, so a file that fails (or is no longer selected) is absent rather than stale.
        # Schematics are cleaned only when the VLM is available (else --no-vlm would wipe what it
        # can't rebuild). Orphans and empty dirs are pruned afterwards.
        targets = [c for c in selected if _only_ok(c)]
        regen_types = {BOM, KICAD} | ({SCHEMATIC} if vlm_available else set())
        if args.only:
            regen_types &= {args.only}  # only wipe the type we'll actually rebuild
        cleaned = clean_outputs(out_dir, regen_types)
        if cleaned:
            print(f"[extract] cleaned {cleaned} existing markdown before full rebuild", file=sys.stderr)
    else:
        # Incremental: regenerate a selected file when its source changed OR it has no markdown yet
        # (e.g. a newly added version folder). Superseded/deleted outputs are removed by the prune
        # step below — so adding a new version correctly resets that product's selection (old
        # version's markdown dropped, new version's generated) without re-running unchanged files.
        targets = [c for c in selected if _only_ok(c) and (c.rel.as_posix() in changed or not _md_path(c).exists())]

    if args.limit:
        targets = targets[: args.limit]

    counts = defaultdict(int)
    skipped_novlm = 0
    errors = []

    for c in targets:
        try:
            page_failures = 0
            if c.doctype == SCHEMATIC:
                if not vlm_available:
                    skipped_novlm += 1
                    continue
                body, page_failures = extract_schematic_pdf(root / c.rel, vlm)
            elif c.doctype == KICAD:
                body = extract_kicad(root / c.rel)
            elif c.rel.suffix.lower() == ".xlsx":
                body = extract_xlsx(root / c.rel)
            else:
                body = extract_csv(root / c.rel)
            write_markdown(out_dir, c, body)
            counts[c.doctype] += 1
            if page_failures:
                # File was written (with placeholders) but flag it so the run fails loudly and CI
                # doesn't commit failure text over good schematics.
                errors.append(f"{c.rel.as_posix()}: {page_failures} schematic page(s) failed VLM extraction")
        except Exception as e:  # noqa: BLE001 - keep going, report at the end
            errors.append(f"{c.rel.as_posix()}: {type(e).__name__}: {e}")

    # Reconcile output against the current selection in BOTH modes: removes markdown whose source
    # was deleted or superseded by a newer version (and cleans empty dirs).
    pruned = prune_orphans(out_dir, selected)

    # ---- summary ----
    sel_by_type = defaultdict(int)
    for c in selected:
        sel_by_type[c.doctype] += 1
    print(f"[extract] mode={mode}  candidates={len(cands)}  selected={len(selected)}", file=sys.stderr)
    print(f"[extract] selected by type: " + ", ".join(f"{k}={v}" for k, v in sorted(sel_by_type.items())), file=sys.stderr)

    # Per-product coverage so missing doc-types are visible at a glance.
    by_product: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    for c in selected:
        by_product[c.product][c.doctype] += 1
    print("[extract] coverage (product: bom/kicad/schematic):", file=sys.stderr)
    for prod in sorted(by_product):
        d = by_product[prod]
        flag = "  <-- no schematic" if not d.get(SCHEMATIC) else ""
        print(f"    {prod}: {d.get(BOM,0)}/{d.get(KICAD,0)}/{d.get(SCHEMATIC,0)}{flag}", file=sys.stderr)
    wrote = ", ".join(f"{k}={v}" for k, v in sorted(counts.items())) or "none"
    print(f"[extract] wrote: {wrote}", file=sys.stderr)
    if skipped_novlm:
        print(f"[extract] schematic PDFs skipped (no VLM / OPENAI_API_KEY): {skipped_novlm}", file=sys.stderr)
    if pruned:
        print(f"[extract] pruned stale markdown: {pruned}", file=sys.stderr)
    if errors:
        print(f"[extract] ERRORS ({len(errors)}):", file=sys.stderr)
        for e in errors:
            print(f"  - {e}", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
